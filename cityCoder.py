# Allows bulk geocoding via google maps api
# Drew Phillips Jan 2018

import openpyxl # to read input excel and write output file
import datetime # to name output file
import requests # to get json from google maps API
import json # to manage json data returned from google maps API
import sys # to get excel sheet name in CLI
import os # to get argv
from cityCoder_pass import *

# TODO: create bounding box for results?

# Enter API key and set up permanent url
url = 'https://maps.googleapis.com/maps/api/geocode/json?'
API_KEY = GITHUB_API_KEY

# Create unique timestamp for output file
run_time = datetime.datetime.now().strftime('%H-%M-%S')

# Main read and write 'module'
print('Opening workbook...')

# Check if user entered excel workbook name in CLI
if (len(sys.argv) < 2):
	print("Remember to enter an xlsx file as an argument value")
	sys.exit()

# user enters workbook name in CLI
workbook_in = sys.argv[1]

# get workbook name without file extension to name output file
workbook_in_name = os.path.splitext(workbook_in)[0]

# Create wb object
wb = openpyxl.load_workbook(workbook_in)
sheet_in = wb.active

# Name new columns lat and long 
# Add as final columns
last_col = sheet_in.max_column
sheet_in.cell(row = 1, column = (last_col + 1)).value = 'lat'
sheet_in.cell(row = 1, column = (last_col + 2)).value = 'long'

# Increase accuracy by passing consistent values to geocoder 
city = ",Madison_Heights"
state = ",MI"

# for all rows except first row
for row_iter in range(2, (sheet_in.max_row+1)):
	
	# Get address string from sheet
	raw_address = sheet_in.cell(row = row_iter, column = 1).value
	
	# Create latlng values by geocoding address
	# 'results' is a class, 'results.text' is a JSON string
	unique_url = url + 'address=%s&key=%s' % ((raw_address + city + state), API_KEY)
	results = requests.get(unique_url)
	
	data = json.loads(results.text)
		
	if (len(data['results'])) == 0:
		print('%i - %s...NOT FOUND' % ((row_iter-1), raw_address))
		pass
		
	else:
		# json is an array with a single object inside: 'data' is a list with a dict inside
		# get lat/lng
		print('%i - %s...OK' % ((row_iter-1), raw_address))
		lat = data['results'][0]['geometry']['location']['lat']
		lng = data['results'][0]['geometry']['location']['lng']
		
		# Add latlng to new column in each row
		sheet_in.cell(row = row_iter, column = (last_col + 1)).value = lat
		sheet_in.cell(row = row_iter, column = (last_col + 2)).value = lng

# Save xlsx as new file with geocoded addresses	
wb.save('%s_geocoded_%s.xlsx' % (workbook_in_name, run_time))
print("Done")